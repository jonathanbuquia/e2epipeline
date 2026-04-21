from __future__ import annotations

import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "excel"
OUTPUT_DIR = ROOT / "public" / "data" / "json"


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return slug or "column"


def unique_keys(headers: list[Any]) -> tuple[list[str], list[dict[str, str]]]:
    counts: dict[str, int] = {}
    keys: list[str] = []
    columns: list[dict[str, str]] = []

    for index, header in enumerate(headers, start=1):
        original = str(header).strip() if header not in (None, "") else f"Column {index}"
        base_key = slugify(original)
        counts[base_key] = counts.get(base_key, 0) + 1
        key = base_key if counts[base_key] == 1 else f"{base_key}_{counts[base_key]}"
        keys.append(key)
        columns.append({"key": key, "label": original})

    return keys, columns


def to_json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def sheet_to_records(sheet) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = list(rows[0])
    keys, columns = unique_keys(headers)
    records: list[dict[str, Any]] = []

    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue

        record = {
            key: to_json_value(row[index]) if index < len(row) else None
            for index, key in enumerate(keys)
        }
        records.append(record)

    return records, columns


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def convert_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook_id = path.stem
    workbook_dir = OUTPUT_DIR / workbook_id
    sheets_meta = []

    for sheet in workbook.worksheets:
        records, columns = sheet_to_records(sheet)
        sheet_file = f"{slugify(sheet.title)}.json"
        write_json(workbook_dir / sheet_file, records)

        sheets_meta.append(
            {
                "name": sheet.title,
                "file": sheet_file,
                "rowCount": len(records),
                "columns": columns,
            }
        )

    metadata = {
        "id": workbook_id,
        "sourceFile": path.name,
        "sheets": sheets_meta,
    }
    write_json(workbook_dir / "metadata.json", metadata)
    return metadata


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbooks = [convert_workbook(path) for path in sorted(SOURCE_DIR.glob("*.xlsx"))]
    write_json(
        OUTPUT_DIR / "index.json",
        {
            "sourceFolder": "excel",
            "workbookCount": len(workbooks),
            "workbooks": workbooks,
        },
    )
    print(f"Converted {len(workbooks)} workbook(s) to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
